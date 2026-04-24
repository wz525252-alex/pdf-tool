# -*- coding: utf-8 -*-
"""
PDF订单统计工具 - Windows GUI版
"""

import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pdfplumber
import openpyxl
import re
import os
from collections import defaultdict
from datetime import datetime


class PDFTool:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF订单统计工具")
        self.root.geometry("600x500")
        self.root.resizable(True, True)

        self.excel_path = tk.StringVar()
        self.pdf_files = []  # [(path, day), ...]

        self.setup_ui()

    def setup_ui(self):
        # 目标表格选择
        frame_excel = ttk.LabelFrame(self.root, text="目标表格")
        frame_excel.pack(fill="x", padx=10, pady=5)

        ttk.Entry(frame_excel, textvariable=self.excel_path, width=50).pack(side="left", padx=5, pady=5)
        ttk.Button(frame_excel, text="选择文件...", command=self.select_excel).pack(side="left", padx=5, pady=5)

        # PDF文件列表
        frame_pdf = ttk.LabelFrame(self.root, text="PDF文件列表")
        frame_pdf.pack(fill="both", expand=True, padx=10, pady=5)

        self.listbox = tk.Listbox(frame_pdf, height=8)
        self.listbox.pack(fill="both", expand=True, padx=5, pady=5)

        frame_btn = ttk.Frame(frame_pdf)
        frame_btn.pack(fill="x", padx=5, pady=5)
        ttk.Button(frame_btn, text="添加文件", command=self.add_pdf).pack(side="left", padx=5)
        ttk.Button(frame_btn, text="清空列表", command=self.clear_pdf).pack(side="left", padx=5)

        # 处理按钮
        frame_action = ttk.Frame(self.root)
        frame_action.pack(fill="x", padx=10, pady=5)
        ttk.Button(frame_action, text="开始处理", command=self.process).pack(pady=5)

        # 日志区域
        frame_log = ttk.LabelFrame(self.root, text="日志")
        frame_log.pack(fill="both", expand=True, padx=10, pady=5)

        self.log_text = tk.Text(frame_log, height=8, state="disabled")
        self.log_text.pack(fill="both", expand=True, padx=5, pady=5)

        scrollbar = ttk.Scrollbar(self.log_text, command=self.log_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_text.config(yscrollcommand=scrollbar.set)

    def log(self, msg):
        self.log_text.config(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def select_excel(self):
        path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )
        if path:
            self.excel_path.set(path)
            self.log(f"已选择表格: {os.path.basename(path)}")

    def add_pdf(self):
        paths = filedialog.askopenfilenames(
            title="选择PDF文件",
            filetypes=[("PDF文件", "*.pdf")]
        )
        for path in paths:
            day = self.extract_day_from_filename(os.path.basename(path))
            self.pdf_files.append((path, day))
            self.listbox.insert("end", f"{os.path.basename(path)}  [日期: {day}号]")
            self.log(f"添加: {os.path.basename(path)} -> {day}号")

    def extract_day_from_filename(self, filename):
        """从文件名提取日期数字，如 '18号单1.pdf' -> 18"""
        match = re.search(r'(\d+)号', filename)
        if match:
            return int(match.group(1))
        # 如果没有"号"，尝试提取开头的数字
        match = re.search(r'^(\d+)', filename)
        if match:
            return int(match.group(1))
        return None

    def clear_pdf(self):
        self.pdf_files.clear()
        self.listbox.delete(0, "end")
        self.log("已清空文件列表")

    def extract_pdf_data(self, pdf_path):
        """从PDF提取订单数据"""
        data_list = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if row and len(row) >= 6:
                            if row[0] in ['序号', '合计', None, ''] or str(row[0]).startswith('合计'):
                                continue
                            try:
                                skc_col = row[2]
                                attr_col = row[4]
                                qty_col = row[5]

                                if skc_col and attr_col and qty_col:
                                    skc_parts = skc_col.split('\n')
                                    product_name = skc_parts[1].strip() if len(skc_parts) >= 2 else skc_col

                                    attr_clean = attr_col.replace('\n', '')
                                    size_match = re.search(r'-([XS|S|M|L|XL]+)-', attr_clean)
                                    if size_match:
                                        size = size_match.group(1)
                                    else:
                                        size_match = re.search(r'-([XS|S|M|L|XL]+)$', attr_clean)
                                        if size_match:
                                            size = size_match.group(1)
                                        else:
                                            continue

                                    if size and product_name:
                                        data_list.append({
                                            'product': product_name,
                                            'size': size,
                                            'qty': int(qty_col)
                                        })
                            except:
                                pass
        return data_list

    def find_column_for_date(self, ws, day):
        """根据日期找到对应的Excel列"""
        base = datetime(1899, 12, 30)
        target_date = datetime(2026, 4, day)
        target_excel = (target_date - base).days

        for col in range(4, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val == target_excel:
                return col
        return None

    def process(self):
        if not self.excel_path.get():
            messagebox.showerror("错误", "请先选择目标Excel文件")
            return

        if not self.pdf_files:
            messagebox.showerror("错误", "请先添加PDF文件")
            return

        self.log("=" * 40)
        self.log("开始处理...")

        try:
            # 按日期分组PDF文件
            pdf_by_day = defaultdict(list)
            for path, day in self.pdf_files:
                if day:
                    pdf_by_day[day].append(path)
                else:
                    self.log(f"警告: 无法识别日期 - {os.path.basename(path)}")

            # 读取Excel
            wb = openpyxl.load_workbook(self.excel_path.get())
            ws = wb['Sheet1']

            # 建立Excel商品索引
            excel_products = {}
            for row_num in range(2, ws.max_row + 1):
                product = ws.cell(row=row_num, column=2).value
                size = ws.cell(row=row_num, column=3).value
                if product and size:
                    excel_products[(product, size)] = row_num

            total_records = 0
            total_qty = 0
            processed_files = []

            # 按日期处理
            for day, pdf_paths in pdf_by_day.items():
                self.log(f"\n处理 {day} 号订单...")

                # 提取并合并数据
                all_data = []
                for pdf_path in pdf_paths:
                    data = self.extract_pdf_data(pdf_path)
                    qty = sum(d['qty'] for d in data)
                    self.log(f"  {os.path.basename(pdf_path)}: {len(data)}条, {qty}件")
                    all_data.extend(data)

                merged_data = defaultdict(int)
                for item in all_data:
                    merged_data[(item['product'], item['size'])] += item['qty']

                # 找到目标列
                target_col = self.find_column_for_date(ws, day)
                if not target_col:
                    self.log(f"  错误: 未找到 {day} 号对应的列")
                    continue

                # 清空目标列
                for row_num in range(2, ws.max_row + 1):
                    ws.cell(row=row_num, column=target_col).value = None

                # 填写数据
                matched = 0
                for (product, size), qty in merged_data.items():
                    key = (product, size)
                    if key in excel_products:
                        row_num = excel_products[key]
                        ws.cell(row=row_num, column=target_col).value = qty
                        matched += 1

                self.log(f"  合并: {len(merged_data)}条, {sum(merged_data.values())}件")
                self.log(f"  匹配: {matched}条 -> 第{target_col}列")

                total_records += len(merged_data)
                total_qty += sum(merged_data.values())
                processed_files.extend(pdf_paths)

            # 保存
            wb.save(self.excel_path.get())
            self.log(f"\n已保存到 {os.path.basename(self.excel_path.get())}")
            self.log(f"总计: {total_records}条, {total_qty}件")

            # 删除已处理的PDF
            for pdf_path in processed_files:
                os.remove(pdf_path)
                self.log(f"已删除: {os.path.basename(pdf_path)}")

            # 清空列表
            self.pdf_files.clear()
            self.listbox.delete(0, "end")

            messagebox.showinfo("完成", f"处理完成!\n总计: {total_records}条, {total_qty}件")

        except Exception as e:
            self.log(f"错误: {str(e)}")
            messagebox.showerror("错误", str(e))


if __name__ == '__main__':
    root = tk.Tk()
    app = PDFTool(root)
    root.mainloop()
